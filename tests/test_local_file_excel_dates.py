"""Pruebas del lector de Excel local (LocalFileConnector) para el caso de
celdas que Excel guardó como FECHA aunque el usuario tipeó un código tipo
"9051-6" (Excel autodetecta "número-número" corto como año-mes y lo
convierte a fecha internamente, aunque se siga viendo igual en la planilla).

Sin esta reconstrucción, ese código se lee como datetime y termina como
"9051-06-01 00:00:00": un SKU irreconocible que no cruza con nada de la
Maestra y crea un producto fantasma (visto en producción con un archivo real
de POEDAGAR: "9051-3", "3076-1", etc. quedaban como filas nuevas con SKU
roto en vez de actualizar el producto que ya existía).
"""
import io

import openpyxl

from backend.connectors.local_file import LocalFileConnector


def _build_xlsx(rows):
    """rows: lista de listas. La primera es el encabezado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _connector(content):
    return LocalFileConnector({"file_content": content, "file_name": "test.xlsx"})


def test_codigo_parecido_a_fecha_no_se_corrompe():
    # Escribir "9051-6" en una celda "General": Excel real hace exactamente
    # esto (autodetecta año-mes y la guarda como datetime), así que se simula
    # con set_explicit_value + number_format en vez de una fecha "genuina".
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nombre", "Código"])
    ws.append(["9051-6 RELOJ", "9051-6"])
    import datetime
    cell = ws.cell(row=3, column=1, value="9051-3 RELOJ")
    cell2 = ws.cell(row=3, column=2, value=datetime.datetime(9051, 3, 1))
    cell2.number_format = "yyyy-m"
    buf = io.BytesIO()
    wb.save(buf)

    records = _connector(buf.getvalue()).fetch_data("Sheet")
    assert records[1]["Código"] == "9051-3"
    assert "00:00:00" not in records[1]["Código"]


def test_stock_y_precio_no_llevan_punto_cero():
    content = _build_xlsx([
        ["Codigo", "Stock", "Precio"],
        ["ABC-1", 10, 182000],
    ])
    records = _connector(content).fetch_data("Sheet")
    assert records[0]["Stock"] == "10"
    assert records[0]["Precio"] == "182000"


def test_sku_normal_no_se_toca():
    content = _build_xlsx([
        ["Codigo", "Nombre"],
        ["365-4", "POEDAGAR CUADRADO METAL DAMA"],
    ])
    records = _connector(content).fetch_data("Sheet")
    assert records[0]["Codigo"] == "365-4"
