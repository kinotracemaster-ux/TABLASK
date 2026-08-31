import os
import io
import re
from datetime import date, datetime
import pandas as pd
import openpyxl
from typing import List, Dict, Any, Tuple
from .base import BaseConnector

# Traduce los tokens de formato de fecha de Excel (yyyy, mm, dd...) a texto real.
# Sin esto, una celda que Excel guardó como FECHA (aunque se vea como "9051-6")
# se lee como datetime y termina como "9051-06-01 00:00:00" — un SKU/código
# irreconocible que no cruza con nada de la Maestra y crea un producto fantasma.
# Usando el number_format de la celda ("yyyy-m") se reconstruye el texto real
# que el usuario ve en Excel ("9051-6"), en vez del ISO completo con hora.
_DATE_TOKEN_RE = re.compile(r'yyyy|yy|mmmm|mmm|mm|m|dddd|ddd|dd|d|hh|h|ss|s', re.IGNORECASE)


def _excel_date_cell_to_str(value, number_format: str) -> str:
    if not isinstance(value, (datetime, date)):
        return "" if value is None else str(value)

    def repl(m):
        tok = m.group(0).lower()
        if tok == 'yyyy':
            return f'{value.year:04d}'
        if tok == 'yy':
            return f'{value.year % 100:02d}'
        if tok == 'mmmm':
            return value.strftime('%B')
        if tok == 'mmm':
            return value.strftime('%b')
        if tok == 'mm':
            return f'{value.month:02d}'
        if tok == 'm':
            return str(value.month)
        if tok == 'dddd':
            return value.strftime('%A')
        if tok == 'ddd':
            return value.strftime('%a')
        if tok == 'dd':
            return f'{value.day:02d}'
        if tok == 'd':
            return str(value.day)
        if tok in ('hh', 'h'):
            hour = getattr(value, 'hour', 0)
            return f'{hour:02d}' if tok == 'hh' else str(hour)
        if tok in ('ss', 's'):
            second = getattr(value, 'second', 0)
            return f'{second:02d}' if tok == 'ss' else str(second)
        return m.group(0)

    fmt = number_format or ''
    if not _DATE_TOKEN_RE.search(fmt):
        # Formato desconocido/no fechable: mejor la fecha completa que nada.
        return value.isoformat(sep=' ')
    return _DATE_TOKEN_RE.sub(repl, fmt)


class LocalFileConnector(BaseConnector):
    """Conector para archivos locales CSV, XLS, XLSX."""

    def __init__(self, connection_config: Dict[str, Any]):
        super().__init__(connection_config)
        self.file_path = self.config.get("file_path")
        # Bytes del archivo guardados en la DB (persisten a los redeploys de
        # Railway, donde el disco es efímero). Si están, se leen de acá.
        self.file_content = self.config.get("file_content")
        # Nombre original (para deducir la extensión aunque leamos de bytes).
        self.file_name = self.config.get("file_name") or self.file_path or ""

    def _buffer(self):
        """Fuente legible por pandas: bytes en memoria si existen, si no la ruta
        en disco. Si no hay ninguno, el archivo se perdió (redeploy) y hay que
        volver a subirlo."""
        if self.file_content:
            return io.BytesIO(self.file_content)
        if self.file_path and os.path.exists(self.file_path):
            return self.file_path
        raise FileNotFoundError(
            "El archivo subido ya no está disponible (probablemente se perdió "
            "tras un redeploy del servidor). Volvé a subirlo para actualizar la fuente."
        )

    def _ext(self) -> str:
        name = (self.file_name or "").lower()
        return name[name.rfind('.'):] if '.' in name else ""

    def fetch_data(self, source_path: str) -> List[Dict[str, Any]]:
        """
        Lee el archivo local.
        source_path es ignorado para CSV, pero usado como nombre de hoja para Excel.
        """
        src = self._buffer()
        ext = self._ext()

        if ext == '.csv':
            df = pd.read_csv(src, dtype=str).fillna("")
            return df.to_dict('records')
        elif ext in ('.xls', '.xlsx'):
            return self._read_excel(src, source_path)
        else:
            raise ValueError("Formato de archivo no soportado. Debe ser csv, xls o xlsx.")

    def _read_excel(self, src, source_path: str) -> List[Dict[str, Any]]:
        """Lee con openpyxl (no pandas) para poder usar el number_format real de
        cada celda: así una celda que Excel guardó como fecha (aunque el usuario
        haya tipeado un código tipo "9051-6") se reconstruye como el texto que
        se ve en Excel, en vez de un datetime con hora pegoteada."""
        wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
        sheet_name = source_path if source_path and source_path != "CSV Data" else wb.sheetnames[0]
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

        rows_iter = ws.iter_rows()
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [("" if c.value is None else str(c.value)) for c in header_row]

        records = []
        for row in rows_iter:
            if all(c.value is None for c in row):
                continue
            values = []
            for c in row:
                if c.data_type == 'd':
                    values.append(_excel_date_cell_to_str(c.value, c.number_format))
                elif c.data_type == 'n' and isinstance(c.value, float) and c.value.is_integer():
                    # Excel guarda todo número como float; sin esto "10" (stock)
                    # o "182000" (precio) salían como "10.0"/"182000.0".
                    values.append(str(int(c.value)))
                else:
                    values.append("" if c.value is None else str(c.value))
            # Completar/recortar para que calce con la cantidad de encabezados.
            values = (values + [""] * len(headers))[:len(headers)]
            records.append(dict(zip(headers, values)))
        return records

    def normalize_data(self, raw_data: List[Dict[str, Any]], field_mappings: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        Aplica los mappings para estandarizar los nombres de las columnas.
        """
        normalized = []
        for row in raw_data:
            new_row = {}
            for src_col, master_col in field_mappings.items():
                new_row[master_col] = str(row.get(src_col, ""))
            normalized.append(new_row)
        return normalized

    def test_connection(self) -> Tuple[bool, str]:
        if self.file_content:
            return True, "Archivo disponible (guardado en la base)."
        if self.file_path and os.path.exists(self.file_path):
            return True, f"Archivo encontrado: {self.file_path}"
        return False, "El archivo subido ya no está disponible: volvé a subirlo."
